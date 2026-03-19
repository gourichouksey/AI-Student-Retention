"""
batch_routes.py
===============
Author: Antra Gupta (@guptaantra6405-creator)
Feature: Excel/CSV batch upload for student risk prediction

Teacher uploads Excel or CSV file with student records.
System predicts dropout risk for every student.
Returns ranked results + downloadable Excel with risk scores.
"""

import io
import os
import sys
import json
import pandas as pd
from pathlib import Path
from flask import Blueprint, request, jsonify, send_file

# Import predictor
sys.path.insert(0, str(Path(__file__).resolve().parent))
from services.predictor import predict_risk

batch_bp = Blueprint("batch", __name__)

# Allowed file extensions
ALLOWED_EXTENSIONS = {"xlsx", "xls", "csv"}

def allowed_file(filename):
    return "." in filename and \
           filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# ─────────────────────────────────────────────
# ENDPOINT 1: Upload file and get predictions
# POST /batch/predict
# ─────────────────────────────────────────────

@batch_bp.route("/predict", methods=["POST"])
def batch_predict():
    """
    Upload Excel or CSV file with student records.
    Returns risk predictions for all students sorted by risk.

    Expected columns (flexible — extra columns ignored):
    student_id, attendance_pct, backlogs, family_income,
    parents_edu, scholarship_status, grade_math,
    grade_science, grade_english
    """
    # Check file was uploaded
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded. Send file as multipart form data with key 'file'"}), 400

    file = request.files["file"]

    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    if not allowed_file(file.filename):
        return jsonify({"error": "Invalid file type. Only .xlsx, .xls, .csv allowed"}), 400

    try:
        # Read file into DataFrame
        filename = file.filename.lower()
        if filename.endswith(".csv"):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file, engine="openpyxl")

        if df.empty:
            return jsonify({"error": "File is empty"}), 400

        if len(df) > 2000:
            return jsonify({"error": "Maximum 2000 students per upload"}), 400

        print(f"[batch] Processing {len(df)} students from {file.filename}")

        # Run prediction for each student
        results = []
        errors  = []

        for idx, row in df.iterrows():
            student_dict = row.to_dict()
            student_id   = student_dict.get("student_id", f"Row_{idx+1}")

            try:
                risk_prob  = predict_risk(student_dict)
                risk_level = get_risk_level(risk_prob)
                top_reason = get_top_reason(student_dict, risk_prob)

                results.append({
                    "student_id":       str(student_id),
                    "risk_probability": round(float(risk_prob), 4),
                    "risk_score":       round(float(risk_prob) * 10, 1),
                    "risk_level":       risk_level,
                    "top_reason":       top_reason,
                    "row_number":       int(idx + 2),  # Excel row (1=header)
                })
            except Exception as e:
                errors.append({
                    "student_id": str(student_id),
                    "error":      str(e)
                })

        # Sort by risk — highest first
        results.sort(key=lambda x: x["risk_probability"], reverse=True)

        # Count by level
        high_count   = sum(1 for r in results if r["risk_level"] == "High")
        medium_count = sum(1 for r in results if r["risk_level"] == "Medium")
        low_count    = sum(1 for r in results if r["risk_level"] == "Low")

        print(f"[batch] Done. High:{high_count} Medium:{medium_count} Low:{low_count}")

        return jsonify({
            "total_students": len(results),
            "high_risk":      high_count,
            "medium_risk":    medium_count,
            "low_risk":       low_count,
            "errors":         len(errors),
            "students":       results,
            "error_details":  errors,
        })

    except Exception as e:
        print(f"[batch] Error processing file: {e}")
        return jsonify({"error": f"Could not process file: {str(e)}"}), 500


# ─────────────────────────────────────────────
# ENDPOINT 2: Download results as Excel
# POST /batch/download
# ─────────────────────────────────────────────

@batch_bp.route("/download", methods=["POST"])
def batch_download():
    """
    Upload file, predict risk, return downloadable Excel
    with risk scores added as new columns.

    Teacher uploads their student Excel →
    gets back same Excel + risk_score + risk_level + top_reason columns added
    """
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]

    if not allowed_file(file.filename):
        return jsonify({"error": "Invalid file type"}), 400

    try:
        # Read file
        filename = file.filename.lower()
        if filename.endswith(".csv"):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file, engine="openpyxl")

        # Predict for each row
        risk_scores  = []
        risk_levels  = []
        top_reasons  = []

        for idx, row in df.iterrows():
            try:
                student_dict = row.to_dict()
                risk_prob    = predict_risk(student_dict)
                risk_scores.append(round(float(risk_prob) * 10, 1))
                risk_levels.append(get_risk_level(risk_prob))
                top_reasons.append(get_top_reason(student_dict, risk_prob))
            except Exception:
                risk_scores.append(None)
                risk_levels.append("Error")
                top_reasons.append("Could not process")

        # Add new columns to original DataFrame
        df["risk_score"]  = risk_scores
        df["risk_level"]  = risk_levels
        df["top_reason"]  = top_reasons

        # Sort by risk score descending
        df = df.sort_values("risk_score", ascending=False)

        # Write to Excel in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Risk Report")

            # Style the Excel — color code by risk level
            workbook  = writer.book
            worksheet = writer.sheets["Risk Report"]

            from openpyxl.styles import PatternFill, Font
            red_fill    = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
            amber_fill  = PatternFill(start_color="FFE0B3", end_color="FFE0B3", fill_type="solid")
            green_fill  = PatternFill(start_color="B3FFB3", end_color="B3FFB3", fill_type="solid")
            bold_font   = Font(bold=True)

            # Style header row
            for cell in worksheet[1]:
                cell.font = bold_font

            # Find risk_level column index
            headers = [cell.value for cell in worksheet[1]]
            try:
                risk_col_idx = headers.index("risk_level") + 1
            except ValueError:
                risk_col_idx = None

            # Color rows based on risk level
            if risk_col_idx:
                for row in worksheet.iter_rows(min_row=2):
                    risk_cell = row[risk_col_idx - 1]
                    if risk_cell.value == "High":
                        for cell in row:
                            cell.fill = red_fill
                    elif risk_cell.value == "Medium":
                        for cell in row:
                            cell.fill = amber_fill
                    elif risk_cell.value == "Low":
                        for cell in row:
                            cell.fill = green_fill

        output.seek(0)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="student_risk_report.xlsx"
        )

    except Exception as e:
        print(f"[batch] Download error: {e}")
        return jsonify({"error": str(e)}), 500


# ─────────────────────────────────────────────
# ENDPOINT 3: Download sample template
# GET /batch/template
# ─────────────────────────────────────────────

@batch_bp.route("/template", methods=["GET"])
def download_template():
    """
    Download a sample Excel template showing the required columns.
    Teacher can fill this in and upload it back.
    """
    sample_data = {
        "student_id":          ["S001", "S002", "S003"],
        "attendance_pct":      [60, 85, 72],
        "backlogs":            [3, 0, 1],
        "family_income":       [80000, 250000, 120000],
        "parents_edu":         ["school", "graduate", "primary"],
        "scholarship_status":  ["No", "Yes", "No"],
        "grade_math":          [45, 78, 62],
        "grade_science":       [50, 82, 58],
        "grade_english":       [55, 75, 65],
    }

    df = pd.DataFrame(sample_data)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Students")
        worksheet = writer.sheets["Students"]
        from openpyxl.styles import Font, PatternFill
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="student_upload_template.xlsx"
    )


# ─────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────

def get_risk_level(prob: float) -> str:
    if prob >= 0.65:
        return "High"
    elif prob >= 0.35:
        return "Medium"
    return "Low"


def get_top_reason(student: dict, risk_prob: float) -> str:
    """Generate a simple human-readable reason based on student data."""
    reasons = []

    att = student.get("attendance_pct", 100)
    if isinstance(att, (int, float)) and att < 65:
        reasons.append("critically low attendance")
    elif isinstance(att, (int, float)) and att < 75:
        reasons.append("low attendance")

    backlogs = student.get("backlogs", 0)
    if isinstance(backlogs, (int, float)) and backlogs >= 2:
        reasons.append(f"{int(backlogs)} subject backlogs")

    income = student.get("family_income", 999999)
    schol  = student.get("scholarship_status", "Yes")
    if isinstance(income, (int, float)) and income < 100000:
        if str(schol).lower() in ["no", "0", "false"]:
            reasons.append("financial stress with no scholarship")

    math    = student.get("grade_math", 100)
    science = student.get("grade_science", 100)
    english = student.get("grade_english", 100)
    failing = sum(
        1 for g in [math, science, english]
        if isinstance(g, (int, float)) and g < 50
    )
    if failing >= 2:
        reasons.append(f"failing {failing} subjects")

    if not reasons:
        if risk_prob >= 0.65:
            return "Multiple combined risk factors"
        elif risk_prob >= 0.35:
            return "Moderate risk — monitor closely"
        return "No major risk factors detected"

    return " + ".join(reasons).capitalize()